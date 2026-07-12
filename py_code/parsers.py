from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pdfplumber import open as pdfopen

from .utils import (
    normalize_invoice_no,
    normalize_key,
    normalize_space,
    period_key_from_date,
    period_key_from_label,
    row_fingerprint,
    safe_float,
)

SECTION_SKIP = {
    "FORM GSTR-2B",
    "FORM SUMMARY - ITC AVAILABLE",
    "PART A",
    "PART B",
    "GSTR-3B",
    "GSTIN OF SUPPLIER",
    "GSTIN OF ISD",
    "GSTIN OF THE SUPPLIER",
    "GSTIN",
    "DESCRIPTION",
    "SR.NO.",
    "S.NO.",
    "SR NO.",
    "NO. OF RECORDS",
    "DOCUMENT TYPE",
    "GOODS AND SERVICES TAX - GSTR-2B",
    "TAXABLE INWARD SUPPLIES RECEIVED FROM REGISTERED PERSONS",
    "AMENDMENTS TO PREVIOUSLY FILED INVOICES BY SUPPLIER",
    "CREDIT WHICH MAY BE AVAILED UNDER FORM GSTR-3B",
}

# Sheets in a GSTR-2B workbook that represent uncommon / attention-worthy
# transaction types, as opposed to the routine B2B feed.
UNCOMMON_SHEETS = {
    "B2B-CDNR", "B2B-CDNRA", "B2B-DNR", "B2B-DNRA",
    "ISD", "ISDA", "ISD(Rejected)", "ISDA(Rejected)",
    "IMPG", "IMPGA", "IMPGSEZ", "IMPGSEZA",
}

NUM = r"[0-9][0-9,]*(?:\.\d+)?"


# ---------------------------------------------------------------------------
# GSTR-1 (outward)
# ---------------------------------------------------------------------------

def parse_gstr1(path: Path) -> dict[str, Any]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        text = _text_from_pdf(path)
        return _parse_gstr1_pdf_text(text, path)
    if ext in {".xlsx", ".xls"}:
        return _parse_gstr1_xlsx(path)
    raise ValueError(f"Unsupported GSTR-1 file: {path}")


def _parse_gstr1_pdf_text(text: str, path: Path) -> dict[str, Any]:
    rows = _tables_flat(path)

    out: dict[str, Any] = {
        "source_file": path.name,
        "return_type": "GSTR-1",
        "period": _first_match(text, r"Tax period\s+([A-Za-z]+)"),
        "financial_year": _first_match(text, r"Financial year\s+([0-9\-]+)"),
        "gstin": _first_match(text, r"\bGSTIN\s+([0-9A-Z]{15})\b"),
        "arn": _first_match(text, r"\bARN\s+([A-Z0-9]+)\b"),
        "arn_date": _first_match(text, r"\bARN date\s+([0-9/]+)\b"),
    }
    out["period_key"] = period_key_from_label(out.get("period", "")) or _fy_month_to_key(
        out.get("financial_year", ""), out.get("period", "")
    )

    # 4A: B2B Regular
    b2b = _find_total_after(rows, "4A - Taxable outward supplies made to registered persons")
    out["b2b_invoice_count"] = _cell_num(b2b[1]) if b2b and len(b2b) > 1 else 0.0
    out["b2b_taxable"] = _cell_num(b2b[3]) if b2b and len(b2b) > 3 else 0.0
    out["b2b_igst"] = _cell_num(b2b[4]) if b2b and len(b2b) > 4 else 0.0
    out["b2b_cgst"] = _cell_num(b2b[5]) if b2b and len(b2b) > 5 else 0.0
    out["b2b_sgst"] = _cell_num(b2b[6]) if b2b and len(b2b) > 6 else 0.0
    out["b2b_cess"] = _cell_num(b2b[7]) if b2b and len(b2b) > 7 else 0.0

    # 4B: B2B reverse charge
    b2brc = _find_total_after(rows, "4B - Taxable outward supplies made to registered persons attracting tax on reverse charge")
    out["b2b_rcm_taxable"] = _cell_num(b2brc[3]) if b2brc and len(b2brc) > 3 else 0.0

    # 5: B2CL
    b2cl = _find_total_after(rows, "5 - Taxable outward inter-state supplies made to unregistered persons")
    out["b2cl_taxable"] = _cell_num(b2cl[3]) if b2cl and len(b2cl) > 3 else 0.0

    # 6A: Exports
    exports = _find_total_after(rows, "6A")
    out["export_taxable"] = _cell_num(exports[3]) if exports and len(exports) > 3 else 0.0

    # 6B: SEZ
    sez = _find_total_after(rows, "6B - Supplies made to SEZ unit or SEZ developer")
    out["sez_count"] = _cell_num(sez[1]) if sez and len(sez) > 1 else 0.0
    out["sez_taxable"] = _cell_num(sez[3]) if sez and len(sez) > 3 else 0.0

    # 6C: Deemed exports
    de = _find_total_after(rows, "6C - Deemed Exports")
    out["deemed_export_taxable"] = _cell_num(de[3]) if de and len(de) > 3 else 0.0

    # 7: B2CS
    b2cs = _find_total_after(rows, "7- Taxable supplies")
    out["b2cs_taxable"] = _cell_num(b2cs[3]) if b2cs and len(b2cs) > 3 else 0.0

    # 9B: Credit/Debit notes (registered), net figure
    cdnr = _find_total_after(rows, "9B - Credit/Debit Notes (Registered)", total_label="Total - Net off debit/credit notes")
    out["cdnr_net_taxable"] = _cell_num(cdnr[3]) if cdnr and len(cdnr) > 3 else 0.0

    # Grand total liability row (last row of the return, not tied to a table header)
    total_row = None
    for row in rows:
        c0 = normalize_space(row[0]) if row and row[0] else ""
        if c0.startswith("Total Liability (Outward supplies"):
            total_row = row
            break
    if total_row:
        out["total_liability"] = _cell_num(total_row[3]) if len(total_row) > 3 else 0.0
        out["total_liability_igst"] = _cell_num(total_row[4]) if len(total_row) > 4 else 0.0
        out["total_liability_cgst"] = _cell_num(total_row[5]) if len(total_row) > 5 else 0.0
        out["total_liability_sgst"] = _cell_num(total_row[6]) if len(total_row) > 6 else 0.0
        out["total_liability_cess"] = _cell_num(total_row[7]) if len(total_row) > 7 else 0.0
    else:
        out["total_liability"] = round(out["b2b_taxable"] + out["sez_taxable"], 2)

    out["outward_total"] = out.get("total_liability", 0.0) or round(
        out["b2b_taxable"] + out["sez_taxable"] + out["export_taxable"]
        + out["b2cl_taxable"] + out["b2cs_taxable"] + out["deemed_export_taxable"], 2
    )
    out["notes"] = "Summary PDF parsed at table level; invoice-level outward lines not available."
    return out


def _parse_gstr1_xlsx(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    flat = _rows_to_flat_dict(rows)
    flat["source_file"] = path.name
    flat["return_type"] = "GSTR-1"
    flat["period_key"] = period_key_from_label(flat.get("period", "") or flat.get("tax period", ""))
    return flat


def _fy_month_to_key(fy: str, month_name: str) -> str:
    """'2025-26' + 'April' -> '2025-04' (handles Jan-Mar rolling into 2nd FY year)."""
    key = period_key_from_label(month_name + " 2000")  # get month number only
    if not key:
        return ""
    mm = key.split("-")[1]
    fy_m = re.match(r"(\d{4})", fy or "")
    if not fy_m:
        return ""
    start_year = int(fy_m.group(1))
    year = start_year if int(mm) >= 4 else start_year + 1
    return f"{year}-{mm}"


# ---------------------------------------------------------------------------
# GSTR-3B (outward liability + ITC)
# ---------------------------------------------------------------------------

def parse_gstr3b(path: Path) -> dict[str, Any]:
    ext = path.suffix.lower()
    if ext == ".pdf":
        text = _text_from_pdf(path)
        return _parse_gstr3b_pdf_text(text, path)
    if ext in {".xlsx", ".xls"}:
        return _parse_gstr3b_xlsx(path)
    raise ValueError(f"Unsupported GSTR-3B file: {path}")


def _parse_payment_table(rows: list[list[str]]) -> dict[str, Any]:
    """Table 6.1 - Payment of tax. For each tax head: liability, ITC utilised
    (with cross-head set-off detail), cash paid, interest, late fee - split
    into (A) Other than reverse charge and (B) Reverse charge / Sec 9(5)."""
    heads = ["igst", "cgst", "sgst", "cess"]
    head_labels = {"Integrated": "igst", "Central": "cgst", "State/UT": "sgst", "Cess": "cess"}

    def _section(start_idx: int) -> dict[str, dict[str, float]]:
        out: dict[str, dict[str, float]] = {}
        j = start_idx
        seen = 0
        while j < len(rows) and seen < 4:
            label = normalize_space(rows[j][0]) if rows[j] else ""
            head = next((v for k, v in head_labels.items() if label.startswith(k)), None)
            if head:
                r = rows[j]
                out[head] = {
                    "tax": _cell_num(r[1]) if len(r) > 1 else 0.0,
                    "itc_igst": _cell_num(r[4]) if len(r) > 4 else 0.0,
                    "itc_cgst": _cell_num(r[5]) if len(r) > 5 else 0.0,
                    "itc_sgst": _cell_num(r[6]) if len(r) > 6 else 0.0,
                    "itc_cess": _cell_num(r[7]) if len(r) > 7 else 0.0,
                    "cash_paid": _cell_num(r[8]) if len(r) > 8 else 0.0,
                    "interest": _cell_num(r[9]) if len(r) > 9 else 0.0,
                    "late_fee": _cell_num(r[10]) if len(r) > 10 else 0.0,
                }
                seen += 1
            j += 1
        for h in heads:
            out.setdefault(h, {"tax": 0.0, "itc_igst": 0.0, "itc_cgst": 0.0, "itc_sgst": 0.0, "itc_cess": 0.0, "cash_paid": 0.0, "interest": 0.0, "late_fee": 0.0})
        return out

    a_idx = next((i for i, r in enumerate(rows) if r and "(A) Other than reverse charge" in normalize_space(r[0])), None)
    b_idx = next((i for i, r in enumerate(rows) if r and "(B) Reverse charge" in normalize_space(r[0])), None)

    return {
        "other_than_rcm": _section(a_idx + 1) if a_idx is not None else {h: {} for h in heads},
        "rcm": _section(b_idx + 1) if b_idx is not None else {h: {} for h in heads},
    }


def _row4(row: list[str] | None, start: int = 1) -> list[float]:
    if not row:
        return [0.0, 0.0, 0.0, 0.0]
    vals = [_cell_num(row[i]) if len(row) > i else 0.0 for i in range(start, start + 4)]
    while len(vals) < 4:
        vals.append(0.0)
    return vals


def _parse_gstr3b_pdf_text(text: str, path: Path) -> dict[str, Any]:
    rows = _tables_flat(path)

    out: dict[str, Any] = {
        "source_file": path.name,
        "return_type": "GSTR-3B",
        "period": _first_match(text, r"\bPeriod\s+([A-Za-z]+)"),
        "financial_year": _first_match(text, r"\bYear\s+([0-9\-]+)"),
        "gstin": _first_match(text, r"\bGSTIN of the supplier\s+([0-9A-Z]{15})\b"),
        "arn": _first_match(text, r"\bARN\s+([A-Z0-9]+)\b"),
        "arn_date": _first_match(text, r"\bDate of ARN\s+([0-9/]+)\b"),
    }
    out["period_key"] = period_key_from_label(out.get("period", "")) or _fy_month_to_key(
        out.get("financial_year", ""), out.get("period", "")
    )

    # 3.1 outward table (values start at col index 1: taxable, igst, cgst, sgst)
    a = _row4(_find_row(rows, "(a) Outward taxable supplies (other than zero rated"))
    out["outward_taxable"], out["outward_igst"], out["outward_cgst"], out["outward_sgst"] = a
    b_row = _find_row(rows, "(b) Outward taxable supplies (zero rated)")
    out["zero_rated_taxable"] = _cell_num(b_row[1]) if b_row and len(b_row) > 1 else 0.0
    d = _row4(_find_row(rows, "(d) Inward supplies (liable to reverse charge)"))
    out["rcm_taxable"], out["rcm_igst"], out["rcm_cgst"], out["rcm_sgst"] = d

    # 4A: Eligible ITC available, by category (values start at col index 1: igst, cgst, sgst, cess)
    out["import_goods_igst"], out["import_goods_cgst"], out["import_goods_sgst"], out["import_goods_cess"] = _row4(
        _find_row(rows, "(1) Import of goods")
    )
    out["import_services_igst"], out["import_services_cgst"], out["import_services_sgst"], out["import_services_cess"] = _row4(
        _find_row(rows, "(2) Import of services")
    )
    out["itc_rcm_igst"], out["itc_rcm_cgst"], out["itc_rcm_sgst"], out["itc_rcm_cess"] = _row4(
        _find_row(rows, "(3) Inward supplies liable to reverse charge (other than 1 & 2 above)")
    )
    out["isd_igst"], out["isd_cgst"], out["isd_sgst"], out["isd_cess"] = _row4(
        _find_row(rows, "(4) Inward supplies from ISD")
    )
    out["all_other_itc_igst"], out["all_other_itc_cgst"], out["all_other_itc_sgst"], out["all_other_itc_cess"] = _row4(
        _find_row(rows, "(5) All other ITC")
    )

    # 4B: ITC reversed
    out["reversal_rule_igst"], out["reversal_rule_cgst"], out["reversal_rule_sgst"], out["reversal_rule_cess"] = _row4(
        _find_row(rows, "(1) As per rules 38,42 & 43 of CGST Rules and section 17(5)")
    )
    out["reversal_other_igst"], out["reversal_other_cgst"], out["reversal_other_sgst"], out["reversal_other_cess"] = _row4(
        _find_row(rows, "(2) Others")
    )

    # 4C: Net ITC available
    out["net_itc_igst"], out["net_itc_cgst"], out["net_itc_sgst"], out["net_itc_cess"] = _row4(
        _find_row(rows, "C. Net ITC available")
    )

    # 4D(2): Ineligible / PoS-restricted ITC
    out["restricted_igst"], out["restricted_cgst"], out["restricted_sgst"], out["restricted_cess"] = _row4(
        _find_row(rows, "(2) Ineligible ITC under section 16(4)")
    )

    # 6.1: Payment of tax - liability, ITC set-off detail, cash, interest, late fee
    out["payment_table"] = _parse_payment_table(rows)

    out["outward_total"] = round(out["outward_taxable"] + out["zero_rated_taxable"], 2)
    out["notes"] = "Summary PDF parsed at table level; invoice-level liability lines not available."
    return out


def _parse_gstr3b_xlsx(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    flat = _rows_to_flat_dict(rows)
    flat["source_file"] = path.name
    flat["return_type"] = "GSTR-3B"
    flat["period_key"] = period_key_from_label(flat.get("period", ""))
    return flat


# ---------------------------------------------------------------------------
# GSTR-2B
# ---------------------------------------------------------------------------

def parse_gstr2b(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    summary = _parse_gstr2b_summary(wb)
    detail = _parse_gstr2b_detail(wb)
    uncommon = [r for r in detail if r.get("sheet") in UNCOMMON_SHEETS]
    rcm_rows = [r for r in detail if normalize_key(r.get("rcm", "")) == "YES"]
    ineligible_rows = [
        r for r in detail
        if r.get("itc_availability") and normalize_key(r.get("itc_availability")) != "YES"
    ]
    period_key = ""
    for r in detail:
        pk = period_key_from_label(r.get("period", ""))
        if pk:
            period_key = pk
            break
    return {
        "source_file": path.name,
        "return_type": "GSTR-2B",
        "period_key": period_key,
        "summary": summary,
        "detail": detail,
        "uncommon": uncommon,
        "rcm_rows": rcm_rows,
        "ineligible_rows": ineligible_rows,
    }


def _parse_gstr2b_summary(wb) -> list[dict[str, Any]]:
    ws = wb["ITC Available"] if "ITC Available" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []
    current_block = ""
    seen = set()

    roman = {"I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}

    for row in rows:
        raw = [normalize_space(v) for v in row]
        if not any(raw):
            continue

        col0 = normalize_key(row[0]) if len(row) > 0 else ""
        col1 = normalize_space(row[1]) if len(row) > 1 else ""
        col2 = normalize_space(row[2]) if len(row) > 2 else ""

        if col0 == "FORM SUMMARY - ITC AVAILABLE" or col0 == "S.NO.":
            continue
        if col0 == "PART A":
            current_block = col1 or "Part A"
            continue
        if col0 in SECTION_SKIP:
            continue

        is_top_level = col0 in roman or col0 in {"PART B"}
        if not is_top_level:
            continue

        numeric = [safe_float(row[i] if len(row) > i else 0) for i in range(3, 7)]
        if not any(v != 0 for v in numeric):
            if not col1 and not col2:
                continue

        key = (col0, normalize_key(col1), normalize_key(col2), *numeric)
        if key in seen:
            continue
        seen.add(key)

        records.append({
            "section": current_block,
            "subsection": col0,
            "heading": col1,
            "gstr3b_table": col2,
            "igst": numeric[0],
            "cgst": numeric[1],
            "sgst": numeric[2],
            "cess": numeric[3],
            "advisory": normalize_space(row[7]) if len(row) > 7 else "",
            "source_sheet": "ITC Available",
        })
    return records


def _find_header_rows(rows):
    for idx, row in enumerate(rows[:10]):
        texts = [normalize_key(v) for v in row if normalize_space(v)]
        joined = " | ".join(texts)
        if ("GSTIN OF SUPPLIER" in joined or "GSTIN OF ISD" in joined or "ICEGATE REFERENCE DATE" in joined or "DOCUMENT NUMBER" in joined):
            header = [normalize_space(v) for v in row]
            return idx, header
    return None


def _parse_gstr2b_detail(wb) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in {"Read me", "ITC Available", "ITC not available", "ITC Reversal", "ITC Rejected"}:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = _find_header_rows(rows)
        if not headers:
            continue
        header_row_idx, header = headers
        for r in rows[header_row_idx + 2:]:
            if not any(v is not None and str(v).strip() != "" for v in r):
                continue
            rec = _parse_detail_row(sheet_name, header, r)
            if rec:
                records.append(rec)
    return records


def _parse_detail_row(sheet_name: str, header: list[Any], row: tuple[Any, ...]) -> dict[str, Any] | None:
    if sheet_name in {"B2B", "B2BA", "B2B-CDNR", "B2B-CDNRA", "B2B (ITC Reversal)", "B2BA (ITC Reversal)", "B2B-DNR", "B2B-DNRA"}:
        return _parse_b2b_like_row(sheet_name, header, row)
    if sheet_name in {"ISD", "ISDA", "ISD(Rejected)", "ISDA(Rejected)"}:
        return _parse_isd_like_row(sheet_name, header, row)
    if sheet_name in {"IMPG", "IMPGA"}:
        return _parse_impg_like_row(sheet_name, header, row)
    if sheet_name in {"IMPGSEZ", "IMPGSEZA"}:
        return _parse_impgsez_like_row(sheet_name, header, row)
    return None


def _parse_b2b_like_row(sheet_name: str, header: list[Any], row: tuple[Any, ...]) -> dict[str, Any] | None:
    cells = [normalize_space(v) for v in row]
    if not cells or not any(cells):
        return None
    gstin = cells[0] if len(cells) > 0 else ""
    inv_no = cells[2] if len(cells) > 2 else ""
    inv_date = cells[4] if len(cells) > 4 else ""
    taxable = safe_float(row[8] if len(row) > 8 else 0)
    igst = safe_float(row[9] if len(row) > 9 else 0)
    cgst = safe_float(row[10] if len(row) > 10 else 0)
    sgst = safe_float(row[11] if len(row) > 11 else 0)
    cess = safe_float(row[12] if len(row) > 12 else 0)

    header_markers = {"GSTIN OF SUPPLIER", "INVOICE NUMBER", "INVOICE DETAILS", "TRADE/LEGAL NAME", "INVOICE TYPE"}
    if normalize_key(gstin) in header_markers or normalize_key(inv_no) in header_markers:
        return None
    if not gstin and not inv_no and taxable == 0 and igst == 0 and cgst == 0 and sgst == 0:
        return None

    return {
        "sheet": sheet_name,
        "gstin_supplier": gstin,
        "party_name": cells[1] if len(cells) > 1 else "",
        "invoice_no": inv_no,
        "invoice_no_norm": normalize_invoice_no(inv_no),
        "invoice_type": cells[3] if len(cells) > 3 else "",
        "invoice_date": inv_date,
        "invoice_value": safe_float(row[5] if len(row) > 5 else 0),
        "place_of_supply": cells[6] if len(cells) > 6 else "",
        "rcm": cells[7] if len(cells) > 7 else "",
        "taxable_value": taxable,
        "igst": igst,
        "cgst": cgst,
        "sgst": sgst,
        "cess": cess,
        "period": cells[13] if len(cells) > 13 else "",
        "filing_date": cells[14] if len(cells) > 14 else "",
        "itc_availability": cells[15] if len(cells) > 15 else "",
        "reason": cells[16] if len(cells) > 16 else "",
        "rate": cells[17] if len(cells) > 17 else "",
        "source": cells[18] if len(cells) > 18 else "",
        "irn": cells[19] if len(cells) > 19 else "",
        "irn_date": cells[20] if len(cells) > 20 else "",
        "fingerprint": row_fingerprint(sheet_name, gstin, inv_no, inv_date, taxable, igst, cgst, sgst, cess),
    }


def _parse_isd_like_row(sheet_name: str, header: list[Any], row: tuple[Any, ...]) -> dict[str, Any] | None:
    cells = [normalize_space(v) for v in row]
    if not any(cells):
        return None
    return {
        "sheet": sheet_name,
        "isd_gstin": cells[0] if len(cells) > 0 else "",
        "party_name": cells[1] if len(cells) > 1 else "",
        "doc_type": cells[2] if len(cells) > 2 else "",
        "doc_no": cells[3] if len(cells) > 3 else "",
        "doc_date": cells[4] if len(cells) > 4 else "",
        "original_invoice_no": cells[5] if len(cells) > 5 else "",
        "original_invoice_date": cells[6] if len(cells) > 6 else "",
        "igst": safe_float(row[7] if len(row) > 7 else 0),
        "cgst": safe_float(row[8] if len(row) > 8 else 0),
        "sgst": safe_float(row[9] if len(row) > 9 else 0),
        "cess": safe_float(row[10] if len(row) > 10 else 0),
        "period": cells[11] if len(cells) > 11 else "",
        "filing_date": cells[12] if len(cells) > 12 else "",
        "eligibility": cells[13] if len(cells) > 13 else "",
        "fingerprint": row_fingerprint(sheet_name, cells[0] if cells else "", cells[3] if len(cells) > 3 else ""),
    }


def _parse_impg_like_row(sheet_name: str, header: list[Any], row: tuple[Any, ...]) -> dict[str, Any] | None:
    cells = [normalize_space(v) for v in row]
    if not any(cells):
        return None
    return {
        "sheet": sheet_name,
        "icegate_ref_date": cells[0] if len(cells) > 0 else "",
        "port_code": cells[1] if len(cells) > 1 else "",
        "boe_no": cells[2] if len(cells) > 2 else "",
        "boe_date": cells[3] if len(cells) > 3 else "",
        "taxable_value": safe_float(row[4] if len(row) > 4 else 0),
        "igst": safe_float(row[5] if len(row) > 5 else 0),
        "cess": safe_float(row[6] if len(row) > 6 else 0),
        "fingerprint": row_fingerprint(sheet_name, cells[2] if len(cells) > 2 else "", cells[3] if len(cells) > 3 else ""),
    }


def _parse_impgsez_like_row(sheet_name: str, header: list[Any], row: tuple[Any, ...]) -> dict[str, Any] | None:
    cells = [normalize_space(v) for v in row]
    if not any(cells):
        return None
    return {
        "sheet": sheet_name,
        "supplier_gstin": cells[0] if len(cells) > 0 else "",
        "party_name": cells[1] if len(cells) > 1 else "",
        "icegate_ref_date": cells[2] if len(cells) > 2 else "",
        "port_code": cells[3] if len(cells) > 3 else "",
        "boe_no": cells[4] if len(cells) > 4 else "",
        "boe_date": cells[5] if len(cells) > 5 else "",
        "taxable_value": safe_float(row[6] if len(row) > 6 else 0),
        "igst": safe_float(row[7] if len(row) > 7 else 0),
        "cess": safe_float(row[8] if len(row) > 8 else 0),
        "fingerprint": row_fingerprint(sheet_name, cells[4] if len(cells) > 4 else "", cells[5] if len(cells) > 5 else ""),
    }


# ---------------------------------------------------------------------------
# Electronic ledgers (Cash / Credit / Liability)
# ---------------------------------------------------------------------------

_LEDGER_LAYOUTS = {
    "electronic_cash_ledger": {
        "fixed_cols": ["sr_no", "date", "time", "reporting_date", "reference_no", "period_label", "description", "txn_type"],
        "tax_heads": ["igst", "cgst", "sgst", "cess"],
        "has_subbreakdown": True,
        "has_balance_group": True,
        "group_has_total_col": False,
    },
    "electronic_liability_ledger": {
        "fixed_cols": ["sr_no", "date", "reference_no", "ledger_used", "description", "txn_type"],
        "tax_heads": ["igst", "cgst", "sgst", "cess"],
        "has_subbreakdown": True,
        "has_balance_group": True,
        "group_has_total_col": False,
    },
    "electronic_credit_ledger": {
        "fixed_cols": ["sr_no", "date", "reference_no", "period_label", "description", "txn_type"],
        "tax_heads": ["igst", "cgst", "sgst", "cess"],
        "has_subbreakdown": False,
        "has_balance_group": True,
        "group_has_total_col": True,
    },
}


def _find_ledger_header_row(rows: list[list[str]]) -> int | None:
    for idx, row in enumerate(rows[:12]):
        if row and normalize_key(row[0]) in {"SR NO", "SR.NO"}:
            return idx
    return None


def parse_ledger_csv(path: Path, ledger_type: str | None = None) -> dict[str, Any]:
    from .utils import detect_file_kind
    ledger_type = ledger_type or detect_file_kind(path)
    layout = _LEDGER_LAYOUTS.get(ledger_type)

    with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
        raw_rows = list(csv.reader(fh))

    if not layout:
        return {
            "source_file": path.name,
            "return_type": ledger_type,
            "transactions": [],
            "opening_balance": {},
            "notes": "Unrecognized ledger type; file skipped.",
            "parse_ok": False,
        }

    header_idx = _find_ledger_header_row(raw_rows)
    if header_idx is None:
        return {
            "source_file": path.name,
            "return_type": ledger_type,
            "transactions": [],
            "opening_balance": {},
            "notes": "Could not locate the 'Sr.No' header row - file layout may have changed on the portal. Re-check this export manually.",
            "parse_ok": False,
        }

    fixed_cols = layout["fixed_cols"]
    n_fixed = len(fixed_cols)
    tax_heads = layout["tax_heads"]
    sub_width = 6 if layout["has_subbreakdown"] else 1
    group_total_col = 1 if layout.get("group_has_total_col") else 0

    data_start = header_idx + 2  # group-header row + sub-header row

    transactions: list[dict[str, Any]] = []
    opening_balance: dict[str, Any] = {}
    skipped_short_rows = 0

    min_len = n_fixed + len(tax_heads) * sub_width * 2 + group_total_col * 2 if layout.get("has_balance_group") \
        else n_fixed + len(tax_heads) * sub_width + group_total_col

    for row in raw_rows[data_start:]:
        if not row or not any(normalize_space(c) for c in row):
            continue
        if len(row) < min_len:
            skipped_short_rows += 1
            continue

        rec: dict[str, Any] = {}
        for i, col_name in enumerate(fixed_cols):
            rec[col_name] = normalize_space(row[i])

        pos = n_fixed
        for head in tax_heads:
            rec[head] = safe_float(row[pos])
            if layout["has_subbreakdown"]:
                rec[f"{head}_total_incl_other"] = safe_float(row[pos + 5])
            pos += sub_width
        pos += group_total_col  # skip the group-level 'Total' column, if present

        if layout.get("has_balance_group"):
            for head in tax_heads:
                rec[f"{head}_balance"] = safe_float(row[pos])
                pos += sub_width
            pos += group_total_col

        is_opening = normalize_key(rec.get("description", "")) == "OPENING BALANCE"
        rec["period_key"] = period_key_from_label(rec.get("period_label", "")) or period_key_from_date(rec.get("date", ""))

        if is_opening:
            opening_balance = {h: rec.get(f"{h}_balance", 0.0) for h in tax_heads}
            opening_balance["as_of_file"] = path.name
            continue

        if not rec.get("date") or rec.get("date") == "-":
            continue

        rec["ledger_type"] = ledger_type
        rec["source_file"] = path.name
        rec["fingerprint"] = row_fingerprint(
            ledger_type, rec.get("reference_no", ""), rec.get("date", ""),
            rec.get("description", ""), rec.get("txn_type", ""),
            *[rec.get(h, 0.0) for h in tax_heads],
        )
        transactions.append(rec)

    notes = "Parsed successfully."
    if skipped_short_rows:
        notes = f"Parsed with {skipped_short_rows} malformed row(s) skipped (column count did not match expected layout)."

    return {
        "source_file": path.name,
        "return_type": ledger_type,
        "transactions": transactions,
        "opening_balance": opening_balance,
        "notes": notes,
        "parse_ok": True,
    }


def merge_ledger_files(parsed_files: list[dict[str, Any]]) -> dict[str, Any]:
    """Combine multiple exports of the same ledger, dropping exact-duplicate
    transactions (e.g. the user re-exporting an overlapping date range),
    while reporting how many duplicates were dropped so nothing is silently
    lost."""
    seen: dict[str, dict[str, Any]] = {}
    duplicates_dropped = 0
    file_notes: list[str] = []
    opening_balances = []

    for pf in parsed_files:
        if not pf.get("parse_ok"):
            file_notes.append(f"{pf['source_file']}: {pf.get('notes')}")
            continue
        if pf.get("opening_balance"):
            opening_balances.append(pf["opening_balance"])
        for tx in pf.get("transactions", []):
            fp = tx["fingerprint"]
            if fp in seen:
                duplicates_dropped += 1
                continue
            seen[fp] = tx

    merged_tx = sorted(seen.values(), key=lambda r: (r.get("period_key", ""), r.get("date", "")))
    return {
        "return_type": parsed_files[0]["return_type"] if parsed_files else "",
        "transactions": merged_tx,
        "opening_balances_seen": opening_balances,
        "duplicates_dropped": duplicates_dropped,
        "file_notes": file_notes,
        "source_files": [pf["source_file"] for pf in parsed_files],
    }


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _text_from_pdf(path: Path) -> str:
    parts: list[str] = []
    with pdfopen(path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            parts.append(txt)
    return "\n".join(parts)


def _tables_flat(path: Path) -> list[list[str]]:
    """Flatten every table on every page into one list of rows (each a list
    of cell strings). Using pdfplumber's table extraction avoids the stray
    whitespace/kerning artifacts that break naive text-regex parsing of
    GST portal PDFs (e.g. '0.00' rendered as '0.0 0')."""
    flat: list[list[str]] = []
    with pdfopen(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    flat.append([c if c is not None else "" for c in row])
    return flat


def _cell_num(cell: str) -> float:
    """Extract a number from a table cell, tolerating stray watermark
    characters pdfplumber sometimes glues onto the text (e.g. 'E\\n143240.00')
    and '-' placeholders for not-applicable amounts."""
    s = normalize_space(cell)
    if not s or s == "-":
        return 0.0
    m = re.search(NUM, s)
    return safe_float(m.group(0)) if m else 0.0


def _find_total_after(
    rows: list[list[str]], header_substr: str, total_label: str = "Total", exclude: str | None = None, lookahead: int = 6
) -> list[str] | None:
    """Find the row whose first cell contains `header_substr` (a section
    header line), then return the next row within `lookahead` rows whose
    first cell starts with `total_label`."""
    for i, row in enumerate(rows):
        c0 = normalize_space(row[0]) if row else ""
        if header_substr in c0 and (not exclude or exclude.lower() not in c0.lower()):
            for j in range(i, min(i + lookahead, len(rows))):
                cj = normalize_space(rows[j][0]) if rows[j] else ""
                if cj.startswith(total_label):
                    return rows[j]
            return None
    return None


def _find_row(rows: list[list[str]], label_substr: str) -> list[str] | None:
    """Find the first row whose first cell contains `label_substr`."""
    for row in rows:
        c0 = normalize_space(row[0]) if row else ""
        if label_substr in c0:
            return row
    return None


def _rows_to_flat_dict(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
    out = {}
    for row in rows:
        vals = [normalize_space(v) for v in row if normalize_space(v)]
        if len(vals) >= 2 and len(vals) <= 6:
            key = normalize_key(vals[0])
            if key and key not in out:
                out[key] = vals[1] if len(vals) > 1 else ""
    return out


def _first_match(text: str, pattern: str) -> str:
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    return normalize_space(m.group(1)) if m else ""
