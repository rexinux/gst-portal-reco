from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import rules

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

NAVY = "1F3864"
LIGHT_BLUE = "DCE6F1"
GREEN = "C6EFCE"
RED = "FFC7CE"
AMBER = "FFEB9C"
GREY = "F2F2F2"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=WHITE)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10)
BOLD_BODY = Font(name="Calibri", size=10, bold=True)
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="808080")

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=GREY)
STATUS_FILL = {
    "Matched": PatternFill("solid", fgColor=GREEN),
    "Mismatch": PatternFill("solid", fgColor=RED),
    "Expected variance": PatternFill("solid", fgColor=AMBER),
    "Informational": PatternFill("solid", fgColor=LIGHT_BLUE),
}
MANUAL_FILL = PatternFill("solid", fgColor="FFF9C4")  # pale yellow - "type here"
GROUP_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY = "#,##0.00"


def _sheet(wb: Workbook, name: str) -> Worksheet:
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    return ws


def _title(ws: Worksheet, text: str, span: int = 8) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    return 3  # next free row


def _section(ws: Worksheet, row: int, text: str, span: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _header_row(ws: Worksheet, row: int, headers: list[str], widths: list[int] | None = None) -> int:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    return row + 1


def _data_row(ws: Worksheet, row: int, values: list[Any], currency_cols: set[int] | None = None, stripe: bool = False, status_col: int | None = None) -> None:
    currency_cols = currency_cols or set()
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.border = BORDER
        c.alignment = Alignment(vertical="center", wrap_text=False)
        if i in currency_cols and isinstance(v, (int, float)):
            c.number_format = CURRENCY
        if stripe and (status_col is None or i != status_col):
            c.fill = STRIPE_FILL
    if status_col:
        status_val = values[status_col - 1]
        fill = STATUS_FILL.get(status_val)
        if fill:
            ws.cell(row=row, column=status_col).fill = fill
            ws.cell(row=row, column=status_col).font = BOLD_BODY


def _note(ws: Worksheet, row: int, text: str, span: int = 8) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return row + 1


def _fmt_period(pk: str) -> str:
    if not pk or pk == "UNKNOWN":
        return "Unknown period"
    try:
        y, m = pk.split("-")
        months = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{months[int(m)]} {y}"
    except Exception:
        return pk


# ---------------------------------------------------------------------------
# Executive Summary
# ---------------------------------------------------------------------------

def _build_executive_summary(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = _sheet(wb, "Executive Summary")
    row = _title(ws, "GST Reconciliation Report - Executive Summary")
    gstin = payload.get("gstin") or "Not detected"
    row = _note(ws, row, f"GSTIN: {gstin}   |   Tolerance applied: Rs. {payload.get('tolerance', 0.5):.2f}   |   "
                          f"Periods covered: {', '.join(_fmt_period(p['period_key']) for p in payload['periods'])}")
    row += 1

    for p in payload["periods"]:
        row = _section(ws, row, f"Period: {_fmt_period(p['period_key'])}")
        g1, g3b, g2b = p.get("gstr1"), p.get("gstr3b"), p.get("gstr2b")

        headers = ["Metric", "Value"]
        row = _header_row(ws, row, headers, widths=[46, 22])
        rows_data = []
        if g1:
            rows_data.append(("GSTR-1 ARN / Date", f"{g1.get('arn','-')}  ({g1.get('arn_date','-')})"))
        if g3b:
            rows_data.append(("GSTR-3B ARN / Date", f"{g3b.get('arn','-')}  ({g3b.get('arn_date','-')})"))
            rows_data.append(("Outward taxable value (3.1(a)+(b))", round(g3b.get("outward_taxable", 0) + g3b.get("zero_rated_taxable", 0), 2)))
            rows_data.append(("Output tax (IGST+CGST+SGST)", round(g3b.get("outward_igst", 0) + g3b.get("outward_cgst", 0) + g3b.get("outward_sgst", 0), 2)))
            rows_data.append(("Net ITC available (4C, IGST+CGST+SGST)", round(g3b.get("net_itc_igst", 0) + g3b.get("net_itc_cgst", 0) + g3b.get("net_itc_sgst", 0), 2)))
            rows_data.append(("Ineligible/restricted ITC (4(D)(2))", round(g3b.get("restricted_cgst", 0) + g3b.get("restricted_sgst", 0) + g3b.get("restricted_igst", 0), 2)))
            rows_data.append(("RCM tax payable (3.1(d))", round(g3b.get("rcm_igst", 0) + g3b.get("rcm_cgst", 0) + g3b.get("rcm_sgst", 0), 2)))
        if g2b:
            rows_data.append(("GSTR-2B invoices (B2B)", len([r for r in g2b.get("detail", []) if r.get("sheet") == "B2B"])))
            rows_data.append(("GSTR-2B uncommon entries (CDNR/DNR/ISD/Import)", len(g2b.get("uncommon", []))))

        outward_rows = p["outward_reco"].get("rows", [])
        itc_rows = p["itc_reco"].get("rows", [])
        matched_o = sum(1 for r in outward_rows if r["status"] == "Matched")
        matched_i = sum(1 for r in itc_rows if r["status"] == "Matched")
        rows_data.append(("Outward reco: matched / total checks", f"{matched_o} / {len(outward_rows)}"))
        rows_data.append(("ITC reco: matched / total checks", f"{matched_i} / {len(itc_rows)}"))
        rows_data.append(("Exception register entries", len(p["exceptions"])))
        rows_data.append(("Duplicate invoice groups flagged", len(p["duplicate_invoices"])))
        if p["missing_returns"]:
            rows_data.append(("Missing returns for this period", ", ".join(p["missing_returns"])))

        for i, (label, val) in enumerate(rows_data):
            currency_cols = {2} if isinstance(val, (int, float)) else set()
            _data_row(ws, row, [label, val], currency_cols=currency_cols, stripe=(i % 2 == 0))
            row += 1
        row += 1

    if payload.get("data_quality_notes"):
        row = _section(ws, row, "Data Quality Flags")
        row = _note(ws, row, f"{len(payload['data_quality_notes'])} item(s) - see the 'Data Quality' sheet for the full list and suggested fix.")
        row += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 26


# ---------------------------------------------------------------------------
# ITC Reconciliation
# ---------------------------------------------------------------------------

def _build_itc_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = _sheet(wb, "ITC Reconciliation")
    row = _title(ws, "ITC Reconciliation - GSTR-3B (4A) vs GSTR-2B")
    row = _note(ws, row, "Compares each ITC bucket in GSTR-3B Table 4(A) against the matching GSTR-2B summary section, tax-head by tax-head. "
                          "'Expected variance' rows are self-assessed RCM tax that GSTR-2B does not carry - not an error.")
    row += 1

    headers = ["Period", "Check", "GSTR-3B (Rs.)", "GSTR-2B (Rs.)", "Difference (Rs.)", "Status", "Remark"]
    row = _header_row(ws, row, headers, widths=[12, 48, 16, 16, 14, 16, 60])

    any_row = False
    for p in payload["periods"]:
        ireco = p["itc_reco"]
        if not ireco.get("available"):
            _data_row(ws, row, [_fmt_period(p["period_key"]), ireco.get("note", "Not available"), "", "", "", "", ""])
            row += 1
            continue
        for i, r in enumerate(ireco["rows"]):
            any_row = True
            v2 = r["value_2"] if r["value_2"] is not None else ""
            diff = r["diff"] if r["diff"] is not None else ""
            _data_row(
                ws, row,
                [_fmt_period(p["period_key"]), r["label"], r["value_1"], v2, diff, r["status"], r.get("remark", "")],
                currency_cols={3, 4, 5}, stripe=(i % 2 == 0), status_col=6,
            )
            row += 1
    if not any_row:
        _note(ws, row, "No ITC reconciliation could be computed - GSTR-3B and/or GSTR-2B were not found for any period.")


# ---------------------------------------------------------------------------
# Exception Register (Annexure)
# ---------------------------------------------------------------------------

def _exception_row_fields(r: dict[str, Any]) -> list[Any]:
    sheet = r.get("sheet", "")
    if sheet in {"B2B"}:
        party = r.get("party_name", "")
        doc = r.get("invoice_no", "")
        doc_date = r.get("invoice_date", "")
        taxable = r.get("taxable_value", 0.0)
        tax = r.get("igst", 0.0) + r.get("cgst", 0.0) + r.get("sgst", 0.0)
        detail = f"RCM: {r.get('rcm','-')}  |  ITC available: {r.get('itc_availability','-')}  |  Reason: {r.get('reason','-')}"
    elif sheet in {"B2B-CDNR", "B2B-CDNRA", "B2B-DNR", "B2B-DNRA"}:
        party = r.get("party_name", "")
        doc = r.get("invoice_no", "")
        doc_date = r.get("invoice_date", "")
        taxable = r.get("taxable_value", 0.0)
        tax = r.get("igst", 0.0) + r.get("cgst", 0.0) + r.get("sgst", 0.0)
        detail = f"{r.get('invoice_type','Note')}  |  ITC available: {r.get('itc_availability','-')}"
    elif sheet in {"ISD", "ISDA", "ISD(Rejected)", "ISDA(Rejected)"}:
        party = r.get("party_name", "")
        doc = r.get("doc_no", "")
        doc_date = r.get("doc_date", "")
        taxable = 0.0
        tax = r.get("igst", 0.0) + r.get("cgst", 0.0) + r.get("sgst", 0.0)
        detail = f"ISD document  |  Eligibility: {r.get('eligibility','-')}"
    elif sheet in {"IMPG", "IMPGA"}:
        party = "Customs (Bill of Entry)"
        doc = r.get("boe_no", "")
        doc_date = r.get("boe_date", "")
        taxable = r.get("taxable_value", 0.0)
        tax = r.get("igst", 0.0)
        detail = f"Port: {r.get('port_code','-')}"
    elif sheet in {"IMPGSEZ", "IMPGSEZA"}:
        party = r.get("party_name", "SEZ supplier")
        doc = r.get("boe_no", "")
        doc_date = r.get("boe_date", "")
        taxable = r.get("taxable_value", 0.0)
        tax = r.get("igst", 0.0)
        detail = f"Port: {r.get('port_code','-')}"
    else:
        party = r.get("party_name", "")
        doc = r.get("invoice_no") or r.get("doc_no") or r.get("boe_no", "")
        doc_date = r.get("invoice_date") or r.get("doc_date") or r.get("boe_date", "")
        taxable = r.get("taxable_value", 0.0)
        tax = r.get("igst", 0.0) + r.get("cgst", 0.0) + r.get("sgst", 0.0)
        detail = ""
    return [sheet, party, doc, doc_date, round(taxable, 2), round(tax, 2), detail]


def _build_exception_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = _sheet(wb, "Exception Register")
    row = _title(ws, "Exception Register - Uncommon GSTR-2B Entries (Annexure)")
    row = _note(ws, row, "Deliberately excludes the routine B2B feed. Only RCM invoices, POS/eligibility-restricted ITC, credit/debit notes, "
                          "ISD credits and import-of-goods entries are listed here - the transactions a CA actually needs to review.")
    row += 1

    headers = ["Period", "Category", "GSTR-2B sheet", "Party / Source", "Document no.", "Document date", "Taxable value (Rs.)", "Tax amount (Rs.)", "Detail"]
    row = _header_row(ws, row, headers, widths=[12, 26, 14, 30, 16, 14, 18, 16, 48])

    any_row = False
    for p in payload["periods"]:
        for i, r in enumerate(p["exceptions"]):
            any_row = True
            fields = _exception_row_fields(r)
            _data_row(
                ws, row,
                [_fmt_period(p["period_key"]), r["category"]] + fields,
                currency_cols={7, 8}, stripe=(i % 2 == 0),
            )
            row += 1
    if not any_row:
        _note(ws, row, "No uncommon entries found - GSTR-2B for the covered period(s) contains only routine B2B invoices.")
        row += 1

    if any(p["duplicate_invoices"] for p in payload["periods"]):
        row += 1
        row = _section(ws, row, "Possible Duplicate Invoices (same supplier GSTIN + invoice number, appears more than once)")
        headers = ["Period", "Supplier GSTIN", "Invoice No.", "Occurrences"]
        row = _header_row(ws, row, headers, widths=[12, 20, 20, 14])
        for p in payload["periods"]:
            for i, d in enumerate(p["duplicate_invoices"]):
                _data_row(ws, row, [_fmt_period(p["period_key"]), d["gstin_supplier"], d["invoice_no"], d["occurrences"]], stripe=(i % 2 == 0))
                row += 1


# ---------------------------------------------------------------------------
# Observations (Auditor's Note)
# ---------------------------------------------------------------------------

def _build_observations_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = _sheet(wb, "Observations")
    row = _title(ws, "Observations & Anomalies - Auditor's Note")
    row = _note(ws, row, "Auto-generated interpretation of the reconciliation results below. Written to be read directly, not as raw data.")
    row += 1

    for p in payload["periods"]:
        row = _section(ws, row, f"Period: {_fmt_period(p['period_key'])}")
        for note in p["observations"]:
            ws.cell(row=row, column=1, value="\u2022")
            c = ws.cell(row=row, column=2, value=note)
            c.font = BODY_FONT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["B"].width = 110


# ---------------------------------------------------------------------------
# Data Quality & Action Items
# ---------------------------------------------------------------------------

def _build_data_quality_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    ws = _sheet(wb, "Data Quality")
    row = _title(ws, "Data Quality & Action Items")
    row = _note(ws, row, "Anything below means part of the reconciliation is incomplete. Fix the input folder and re-run to close the gap.")
    row += 1

    headers = ["#", "Issue"]
    row = _header_row(ws, row, headers, widths=[6, 120])
    notes = payload.get("data_quality_notes", [])
    if not notes:
        _note(ws, row, "No data quality issues detected - all expected files were found and parsed cleanly.")
        return
    for i, n in enumerate(notes, start=1):
        _data_row(ws, row, [i, n], stripe=(i % 2 == 0))
        ws.row_dimensions[row].height = 28
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        row += 1


# ---------------------------------------------------------------------------
# Reason code reference (only codes actually used this run)
# ---------------------------------------------------------------------------

def _build_reference_sheet(wb: Workbook, payload: dict[str, Any]) -> None:
    used_codes = set()
    for p in payload["periods"]:
        for r in p["outward_reco"].get("rows", []):
            used_codes.add(r.get("reason_code"))
        for r in p["itc_reco"].get("rows", []):
            used_codes.add(r.get("reason_code"))
    used_codes.discard(None)
    used_codes.discard("MATCHED")
    if not used_codes:
        return

    ws = _sheet(wb, "Reason Codes")
    row = _title(ws, "Reason Code Reference")
    row = _note(ws, row, "Only the codes that were actually triggered in this report are listed.")
    row += 1
    headers = ["Code", "Title", "Explanation", "Suggested action", "Severity"]
    row = _header_row(ws, row, headers, widths=[22, 26, 60, 40, 12])
    catalog = {c.code: c for c in rules.mismatch_catalog()}
    for i, code in enumerate(sorted(used_codes)):
        rc = catalog.get(code)
        if not rc:
            continue
        _data_row(ws, row, [rc.code, rc.title, rc.explanation, rc.action, rc.severity], stripe=(i % 2 == 0))
        ws.row_dimensions[row].height = 45
        for col in (3, 4):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="center")
        row += 1


# ---------------------------------------------------------------------------
# Processed Data - the single engine sheet. One row per FY month, fixed
# order (Apr..Mar), so every formula elsewhere in the workbook can reference
# it by a direct cell address instead of a lookup. Edit a number here and
# every sheet that depends on it recalculates automatically.
# ---------------------------------------------------------------------------

PROCESSED_DATA_SPEC: list[tuple[str, list[tuple[str, str]]]] = [
    ("GSTR-1 (Outward)", [
        ("g1_taxable", "Taxable Value"), ("g1_igst", "IGST"), ("g1_cgst", "CGST"), ("g1_sgst", "SGST"),
        ("g1_arn", "ARN"), ("g1_arn_date", "ARN Date"),
    ]),
    ("GSTR-3B (Outward)", [
        ("g3b_taxable", "Taxable Value"), ("g3b_igst", "IGST"), ("g3b_cgst", "CGST"), ("g3b_sgst", "SGST"),
        ("g3b_arn", "ARN"), ("g3b_arn_date", "ARN Date"),
    ]),
    ("GSTR-3B Net ITC (4C)", [
        ("itc_igst", "IGST"), ("itc_cgst", "CGST"), ("itc_sgst", "SGST"),
    ]),
    ("GSTR-3B RCM Liability (3.1(d))", [
        ("rcm_taxable", "Taxable Value"), ("rcm_igst", "IGST"), ("rcm_cgst", "CGST"), ("rcm_sgst", "SGST"),
    ]),
    ("Payment 6.1 - IGST (Other than RCM)", [
        ("pay_igst_liability", "Liability"), ("pay_igst_itc_same", "Paid via IGST ITC"),
        ("pay_igst_itc_cross", "Paid via CGST/SGST ITC"), ("pay_igst_cash", "Paid in Cash"),
        ("pay_igst_interest", "Interest Paid"), ("pay_igst_latefee", "Late Fee Paid"),
    ]),
    ("Payment 6.1 - CGST (Other than RCM)", [
        ("pay_cgst_liability", "Liability"), ("pay_cgst_itc_same", "Paid via CGST ITC"),
        ("pay_cgst_itc_cross", "Paid via IGST ITC"), ("pay_cgst_cash", "Paid in Cash"),
        ("pay_cgst_interest", "Interest Paid"), ("pay_cgst_latefee", "Late Fee Paid"),
    ]),
    ("Payment 6.1 - SGST (Other than RCM)", [
        ("pay_sgst_liability", "Liability"), ("pay_sgst_itc_same", "Paid via SGST ITC"),
        ("pay_sgst_itc_cross", "Paid via IGST ITC"), ("pay_sgst_cash", "Paid in Cash"),
        ("pay_sgst_interest", "Interest Paid"), ("pay_sgst_latefee", "Late Fee Paid"),
    ]),
    ("Payment 6.1 - RCM (always cash)", [
        ("rcm_pay_igst_cash", "IGST Paid"), ("rcm_pay_cgst_cash", "CGST Paid"), ("rcm_pay_sgst_cash", "SGST Paid"),
        ("rcm_pay_interest", "Interest Paid"), ("rcm_pay_latefee", "Late Fee Paid"),
    ]),
    ("GSTR-2B", [
        ("g2b_avail_igst", "ITC Available IGST"), ("g2b_avail_cgst", "ITC Available CGST"), ("g2b_avail_sgst", "ITC Available SGST"),
        ("g2b_uncommon_count", "Uncommon Entries"), ("g2b_ineligible_cgst", "Ineligible ITC CGST"), ("g2b_ineligible_sgst", "Ineligible ITC SGST"),
    ]),
    ("Ledgers", [
        ("cash_credited", "Cash Ledger Credited"), ("cash_debited", "Cash Ledger Debited"),
        ("credit_accrued", "Credit Ledger Accrued"), ("credit_utilised", "Credit Ledger Utilised"), ("credit_carried_fwd", "Credit Carried Fwd"),
    ]),
]

TEXT_FIELDS = {"g1_arn", "g1_arn_date", "g3b_arn", "g3b_arn_date"}


def _build_processed_data_sheet(wb: Workbook, payload: dict[str, Any]) -> dict[str, Any]:
    ws = _sheet(wb, "Processed Data")
    ws.sheet_properties.tabColor = NAVY

    col = 2  # column A is Period
    field_col: dict[str, str] = {}
    for group_label, fields in PROCESSED_DATA_SPEC:
        start_col = col
        for field_key, _ in fields:
            field_col[field_key] = get_column_letter(col)
            col += 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)
        gc = ws.cell(row=1, column=start_col, value=group_label)
        gc.font = HEADER_FONT
        gc.fill = HEADER_FILL
        gc.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="")
    ws.cell(row=2, column=1, value="Period").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 10

    col = 2
    for _, fields in PROCESSED_DATA_SPEC:
        for field_key, header_label in fields:
            c = ws.cell(row=2, column=col, value=header_label)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 15 if field_key not in TEXT_FIELDS else 16
            col += 1
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "B3"

    data_start_row = 3
    processed = payload["processed_data"]
    for i, rec in enumerate(processed):
        r = data_start_row + i
        ws.cell(row=r, column=1, value=rec["label"]).font = BOLD_BODY
        col = 2
        for _, fields in PROCESSED_DATA_SPEC:
            for field_key, _ in fields:
                val = rec.get(field_key, "" if field_key in TEXT_FIELDS else 0.0)
                c = ws.cell(row=r, column=col, value=val)
                c.font = BODY_FONT
                if field_key not in TEXT_FIELDS:
                    c.number_format = CURRENCY if "count" not in field_key else "0"
                if i % 2 == 0:
                    c.fill = STRIPE_FILL
                col += 1

    data_end_row = data_start_row + len(processed) - 1
    ws.sheet_view.showGridLines = True
    return {"sheet": "Processed Data", "cols": field_col, "start": data_start_row, "end": data_end_row}


def _ref(pd_map: dict[str, Any], field_key: str, row_offset: int) -> str:
    """Build a formula reference into Processed Data for the given field and
    0-indexed month offset (0=first FY month)."""
    col = pd_map["cols"][field_key]
    row = pd_map["start"] + row_offset
    return f"'{pd_map['sheet']}'!{col}{row}"


# ---------------------------------------------------------------------------
# Reconciliation Grid - the sheet requested: GSTR-3B Sales/ITC/RCM/Cash-paid,
# GSTR-1 Sales, a live GSTR-1-vs-GSTR-3B diff (so the same outward figure
# doesn't have to be read twice and mentally compared), and a Tally
# comparison block (manual entry, exactly like the reference template).
# ---------------------------------------------------------------------------

def _grid_header_block(ws: Worksheet, row: int, group_label: str, sub_headers: list[str], start_col: int) -> int:
    end_col = start_col + len(sub_headers) - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    gc = ws.cell(row=row, column=start_col, value=group_label)
    gc.font = HEADER_FONT
    gc.fill = HEADER_FILL
    gc.alignment = Alignment(horizontal="center", vertical="center")
    for i, sh in enumerate(sub_headers):
        c = ws.cell(row=row + 1, column=start_col + i, value=sh)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(start_col + i)].width = 13
    return end_col + 1


def _build_reconciliation_grid_sheet(wb: Workbook, payload: dict[str, Any], pd_map: dict[str, Any]) -> None:
    ws = _sheet(wb, "Reconciliation Grid")
    row = _title(ws, "Monthly Reconciliation Grid - GSTR-3B, GSTR-1 and Books (Tally)", span=30)
    row = _note(ws, row, "Every number below is a live formula pointing at the 'Processed Data' sheet - edit a figure there (or re-run the tool) and this grid recalculates. "
                          "Tally cells (yellow) are the only ones you type into directly.", span=30)
    row += 1

    header_row = row
    ws.cell(row=header_row, column=1, value="")
    col = 2
    blocks = [
        ("GSTR-3B Sales", ["Taxable", "IGST", "CGST", "SGST"], ["g3b_taxable", "g3b_igst", "g3b_cgst", "g3b_sgst"]),
        ("GSTR-3B Net ITC", ["IGST", "CGST", "SGST"], ["itc_igst", "itc_cgst", "itc_sgst"]),
        ("GSTR-3B RCM Liability", ["Taxable", "IGST", "CGST", "SGST"], ["rcm_taxable", "rcm_igst", "rcm_cgst", "rcm_sgst"]),
        ("Tax Paid in Cash (Regular)", ["IGST", "CGST", "SGST"], ["pay_igst_cash", "pay_cgst_cash", "pay_sgst_cash"]),
        ("Tax Paid in Cash (RCM)", ["IGST", "CGST", "SGST"], ["rcm_pay_igst_cash", "rcm_pay_cgst_cash", "rcm_pay_sgst_cash"]),
        ("GSTR-1 Sales", ["Taxable", "IGST", "CGST", "SGST"], ["g1_taxable", "g1_igst", "g1_cgst", "g1_sgst"]),
    ]
    col_positions: dict[str, list[int]] = {}
    for label, subs, keys in blocks:
        start_col = col
        col = _grid_header_block(ws, header_row, label, subs, start_col)
        col_positions[label] = list(range(start_col, col))

    diff_start = col
    diff_subs = ["Taxable", "IGST", "CGST", "SGST"]
    col = _grid_header_block(ws, header_row, "GSTR-1 vs GSTR-3B (Sales already reported once - this is the check, not a third figure)", diff_subs, diff_start)
    ws.row_dimensions[header_row].height = 34
    ws.row_dimensions[header_row + 1].height = 30
    ws.freeze_panes = ws.cell(row=header_row + 2, column=2)

    data_row0 = header_row + 2
    n_months = len(payload["processed_data"])
    for i in range(n_months):
        r = data_row0 + i
        ws.cell(row=r, column=1, value=payload["processed_data"][i]["label"]).font = BOLD_BODY
        if i % 2 == 0:
            ws.cell(row=r, column=1).fill = STRIPE_FILL

        for label, subs, keys in blocks:
            for j, key in enumerate(keys):
                c_idx = col_positions[label][j]
                c = ws.cell(row=r, column=c_idx, value=f"={_ref(pd_map, key, i)}")
                c.number_format = CURRENCY
                c.font = BODY_FONT
                c.border = BORDER
                if i % 2 == 0:
                    c.fill = STRIPE_FILL

        g1_cols = col_positions["GSTR-1 Sales"]
        g3_cols = col_positions["GSTR-3B Sales"]
        for j in range(4):
            c = ws.cell(row=r, column=diff_start + j, value=f"={get_column_letter(g1_cols[j])}{r}-{get_column_letter(g3_cols[j])}{r}")
            c.number_format = CURRENCY
            c.font = BODY_FONT
            c.border = BORDER
            if i % 2 == 0:
                c.fill = STRIPE_FILL

    total_row = data_row0 + n_months + 1
    ws.cell(row=total_row, column=1, value="Total").font = BOLD_BODY
    for c_idx in range(2, diff_start + 4):
        col_letter = get_column_letter(c_idx)
        c = ws.cell(row=total_row, column=c_idx, value=f"=SUM({col_letter}{data_row0}:{col_letter}{data_row0 + n_months - 1})")
        c.number_format = CURRENCY
        c.font = BOLD_BODY
        c.fill = GROUP_FILL

    diff_range = f"{get_column_letter(diff_start)}{data_row0}:{get_column_letter(diff_start + 3)}{data_row0 + n_months - 1}"
    ws.conditional_formatting.add(
        diff_range,
        FormulaRule(formula=[f"ABS({get_column_letter(diff_start)}{data_row0})<=0.5"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        diff_range,
        FormulaRule(formula=[f"ABS({get_column_letter(diff_start)}{data_row0})>0.5"], fill=PatternFill("solid", fgColor=RED)),
    )

    # --- Tally comparison block ---
    tally_title_row = total_row + 3
    row = _section(ws, tally_title_row, "Books (Tally) Comparison - type your Tally sales/purchase figures below; the difference vs GSTR-3B is automatic", span=30)
    header_row2 = row + 1
    col = 2
    tally_blocks = [
        ("Tally Sales (type here)", ["Taxable", "IGST", "CGST", "SGST"]),
        ("Tally Purchase / ITC (type here)", ["IGST", "CGST", "SGST"]),
    ]
    tcol_positions: dict[str, list[int]] = {}
    for label, subs in tally_blocks:
        start_col = col
        col = _grid_header_block(ws, header_row2, label, subs, start_col)
        tcol_positions[label] = list(range(start_col, col))

    diff2_start = col
    col = _grid_header_block(ws, header_row2, "Diff: Tally Sales vs GSTR-3B Sales", ["Taxable", "IGST", "CGST", "SGST"], diff2_start)
    diff3_start = col
    col = _grid_header_block(ws, header_row2, "Diff: Tally Purchase vs GSTR-3B Net ITC", ["IGST", "CGST", "SGST"], diff3_start)
    ws.row_dimensions[header_row2].height = 34
    ws.row_dimensions[header_row2 + 1].height = 30

    tdata_row0 = header_row2 + 2
    for i in range(n_months):
        r = tdata_row0 + i
        ws.cell(row=r, column=1, value=payload["processed_data"][i]["label"]).font = BOLD_BODY
        for c_idx in tcol_positions["Tally Sales (type here)"] + tcol_positions["Tally Purchase / ITC (type here)"]:
            c = ws.cell(row=r, column=c_idx, value=0)
            c.number_format = CURRENCY
            c.fill = MANUAL_FILL
            c.border = BORDER

        ts_cols = tcol_positions["Tally Sales (type here)"]
        g3_cols = col_positions["GSTR-3B Sales"]
        for j in range(4):
            c = ws.cell(row=r, column=diff2_start + j,
                        value=f"={get_column_letter(ts_cols[j])}{r}-{get_column_letter(g3_cols[j])}{data_row0 + i}")
            c.number_format = CURRENCY
            c.border = BORDER

        tp_cols = tcol_positions["Tally Purchase / ITC (type here)"]
        itc_cols = col_positions["GSTR-3B Net ITC"]
        for j in range(3):
            c = ws.cell(row=r, column=diff3_start + j,
                        value=f"={get_column_letter(tp_cols[j])}{r}-{get_column_letter(itc_cols[j])}{data_row0 + i}")
            c.number_format = CURRENCY
            c.border = BORDER

    ttotal_row = tdata_row0 + n_months + 1
    ws.cell(row=ttotal_row, column=1, value="Total").font = BOLD_BODY
    for c_idx in range(2, diff3_start + 3):
        col_letter = get_column_letter(c_idx)
        c = ws.cell(row=ttotal_row, column=c_idx, value=f"=SUM({col_letter}{tdata_row0}:{col_letter}{tdata_row0 + n_months - 1})")
        c.number_format = CURRENCY
        c.font = BOLD_BODY
        c.fill = GROUP_FILL


# ---------------------------------------------------------------------------
# Payment & Compliance Detail - mines GSTR-3B Table 6.1 fully: for each tax
# head, how much liability was cleared via same-head ITC, cross-head ITC,
# and cash, plus interest and late fee actually recorded on the return.
# ---------------------------------------------------------------------------

def _build_payment_detail_sheet(wb: Workbook, payload: dict[str, Any], pd_map: dict[str, Any]) -> None:
    ws = _sheet(wb, "Payment Detail")
    row = _title(ws, "Payment & Compliance Detail - GSTR-3B Table 6.1", span=24)
    row = _note(ws, row, "How each month's liability was actually cleared - same-head ITC, cross-utilised ITC from another head, cash, and any interest/late fee "
                          "recorded on the return itself (not inferred from the ledger). Live formulas from Processed Data.", span=24)
    row += 1

    header_row = row
    col = 2
    blocks = [
        ("IGST", ["Liability", "Via IGST ITC", "Via CGST/SGST ITC", "Cash Paid", "Interest", "Late Fee"],
         ["pay_igst_liability", "pay_igst_itc_same", "pay_igst_itc_cross", "pay_igst_cash", "pay_igst_interest", "pay_igst_latefee"]),
        ("CGST", ["Liability", "Via CGST ITC", "Via IGST ITC", "Cash Paid", "Interest", "Late Fee"],
         ["pay_cgst_liability", "pay_cgst_itc_same", "pay_cgst_itc_cross", "pay_cgst_cash", "pay_cgst_interest", "pay_cgst_latefee"]),
        ("SGST", ["Liability", "Via SGST ITC", "Via IGST ITC", "Cash Paid", "Interest", "Late Fee"],
         ["pay_sgst_liability", "pay_sgst_itc_same", "pay_sgst_itc_cross", "pay_sgst_cash", "pay_sgst_interest", "pay_sgst_latefee"]),
        ("RCM (always cash)", ["IGST Paid", "CGST Paid", "SGST Paid", "Interest", "Late Fee"],
         ["rcm_pay_igst_cash", "rcm_pay_cgst_cash", "rcm_pay_sgst_cash", "rcm_pay_interest", "rcm_pay_latefee"]),
    ]
    col_positions: dict[str, list[int]] = {}
    for label, subs, keys in blocks:
        start_col = col
        col = _grid_header_block(ws, header_row, label, subs, start_col)
        col_positions[label] = list(range(start_col, col))
    ws.row_dimensions[header_row].height = 24
    ws.row_dimensions[header_row + 1].height = 30
    ws.freeze_panes = ws.cell(row=header_row + 2, column=2)

    data_row0 = header_row + 2
    n_months = len(payload["processed_data"])
    interest_latefee_cols = []
    for label, subs, keys in blocks:
        for j, sh in enumerate(subs):
            if sh in ("Interest", "Late Fee"):
                interest_latefee_cols.append(col_positions[label][j])

    for i in range(n_months):
        r = data_row0 + i
        ws.cell(row=r, column=1, value=payload["processed_data"][i]["label"]).font = BOLD_BODY
        if i % 2 == 0:
            ws.cell(row=r, column=1).fill = STRIPE_FILL
        for label, subs, keys in blocks:
            for j, key in enumerate(keys):
                c_idx = col_positions[label][j]
                c = ws.cell(row=r, column=c_idx, value=f"={_ref(pd_map, key, i)}")
                c.number_format = CURRENCY
                c.font = BODY_FONT
                c.border = BORDER
                if i % 2 == 0:
                    c.fill = STRIPE_FILL

    total_row = data_row0 + n_months + 1
    ws.cell(row=total_row, column=1, value="Total").font = BOLD_BODY
    last_col = col - 1
    for c_idx in range(2, last_col + 1):
        col_letter = get_column_letter(c_idx)
        c = ws.cell(row=total_row, column=c_idx, value=f"=SUM({col_letter}{data_row0}:{col_letter}{data_row0 + n_months - 1})")
        c.number_format = CURRENCY
        c.font = BOLD_BODY
        c.fill = GROUP_FILL
        if c_idx in interest_latefee_cols:
            c.fill = PatternFill("solid", fgColor=AMBER)

    row = total_row + 2
    row = _note(ws, row, "Interest/Late Fee columns highlighted amber in the Total row - any non-zero total there is worth a closer look at which month triggered it.", span=24)


# ---------------------------------------------------------------------------
# Hidden raw sheets - full transaction-level detail kept for audit trail /
# traceability, but hidden by default so the workbook opens on the readable
# summary sheets, not a wall of routine invoice rows.
# ---------------------------------------------------------------------------

def _build_raw_table_sheet(wb: Workbook, name: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    if not rows:
        return
    ws = _sheet(wb, name)
    for i, col_name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=col_name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for r_idx, rec in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            val = rec.get(col_name, "")
            if isinstance(val, float):
                val = round(val, 2)
            ws.cell(row=r_idx, column=c_idx, value=val)
    ws.freeze_panes = "A2"
    ws.sheet_state = "hidden"


def _build_raw_sheets(wb: Workbook, payload: dict[str, Any]) -> None:
    b2b_rows = []
    for pk, g2b in payload.get("gstr2b_by_period", {}).items():
        for r in g2b.get("detail", []):
            if r.get("sheet") == "B2B":
                b2b_rows.append({**r, "period_key": pk})
    _build_raw_table_sheet(
        wb, "Raw - GSTR2B B2B", b2b_rows,
        ["period_key", "gstin_supplier", "party_name", "invoice_no", "invoice_date", "taxable_value",
         "igst", "cgst", "sgst", "cess", "rcm", "itc_availability", "reason"],
    )
    _build_raw_table_sheet(
        wb, "Raw - Cash Ledger", payload["merged_cash"].get("transactions", []),
        ["period_key", "date", "reference_no", "description", "txn_type", "igst", "cgst", "sgst", "cess", "source_file"],
    )
    _build_raw_table_sheet(
        wb, "Raw - Credit Ledger", payload["merged_credit"].get("transactions", []),
        ["period_key", "date", "reference_no", "description", "txn_type", "igst", "cgst", "sgst", "cess", "source_file"],
    )
    _build_raw_table_sheet(
        wb, "Raw - Liability Ledger", payload["merged_liability"].get("transactions", []),
        ["period_key", "date", "reference_no", "description", "txn_type", "igst", "cgst", "sgst", "cess", "source_file"],
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_report(payload: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _build_executive_summary(wb, payload)
    pd_map = _build_processed_data_sheet(wb, payload)
    _build_reconciliation_grid_sheet(wb, payload, pd_map)
    _build_payment_detail_sheet(wb, payload, pd_map)
    _build_observations_sheet(wb, payload)
    _build_itc_sheet(wb, payload)
    _build_exception_sheet(wb, payload)
    _build_data_quality_sheet(wb, payload)
    _build_reference_sheet(wb, payload)
    _build_raw_sheets(wb, payload)

    wb.save(output_path)
